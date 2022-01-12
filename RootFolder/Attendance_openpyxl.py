# importing libraries 
import cv2
import face_recognition
import numpy as np
import os
from datetime import date,datetime
from openpyxl import workbook,load_workbook,styles
from openpyxl.styles import Alignment, colors,Font,PatternFill,Border,Side
import csv

# Defining colors 
Black = "00000000"
orange = "00F68B27"
Blue = "FC766AFF"
Light_Black = "004B4C4E"
Light_Blue = "004A8DDC"
pink = "00FF00FF"
light_red = "00E77577"
green = "00008000"
light_green = "00A6CF4A"
white = "FFFFFFFF"

# path of the image 
path_img= "C:\\Users\\Administrator\\Documents\\FaceRecognitionProject\\ImagesAttendance"

# counting the number of files
files = os.listdir(path_img)
file_count = len(files) # counting the number of files 

images = [] #To store the image names 
classNames = [] # to store the name of the image excluding the extension
myList = os.listdir(path_img)
path = "C:\\Users\\Administrator\\Documents\\FaceRecognitionProject\\RootFolder\\"
wb = load_workbook(path + "Attendance_Final.xlsx") # loading the workbook
ws = wb.active

for cl in myList:
    curImg = cv2.imread(f'{path_img}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])

 # function which finds the encoding of the image
def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList
            

#This function setups the Name and Roll Number of the student in the excel file    
def Setup():

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.merge_cells("A1:A2")
    cell = ws.cell(row = 1,column=1)
    cell.value = "Name"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(name = "Arial Black",size =11,color=Black)
    ws.merge_cells("B1:B2")
    cell = ws.cell(row = 1,column = 2)
    cell.value = "Roll NO"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(name = "Arial Black",size =11,color=Black)
    myList = os.listdir(path_img)
    val = 3
    j = 1
    for Names in myList:
        classNames.append(os.path.splitext(Names)[0])
        cel = f'A{str(val)}'
        cel_2 = f'B{str(val)}'
        ws[cel].value  = classNames[val-3]
        ws[cel].font = Font(name="Rockwell", size=13, color=Blue)
        ws[cel_2].value = j
        ws[cel_2].alignment = Alignment(horizontal='center', vertical='center')
        ws[cel_2].font = Font(name = "Arial Black",size =13,color=orange)
        val+=1
        j+=1
    

#This function Creates a new register every time the date changes 
def Create(cell1, cell2,col): 
    concat1 = cell1 
    concat2 = cell2  
    ws.column_dimensions[cell1].width = 20
    ws.column_dimensions[cell2].width = 20
    ws.merge_cells(f"{concat1}1:{concat2}1")
    cell = ws.cell(row = 1,column = col)
    today = date.today()
    d1 = today.strftime("%d/%m/%y")
    cell.value = f"Date = {d1}"
    # cell.value = "Date = 09/01/22"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(name = "Arial Black",size =11,color=Light_Black)
    ws[f"{concat1}2"].value = "Arrival Time"
    ws[f"{concat2}2"].value = "Status"
    ws[f"{concat1}2"].alignment = Alignment(horizontal='center', vertical='center')
    ws[f"{concat2}2"].alignment = Alignment(horizontal='center', vertical='center')
    ws[f"{concat1}2"].font = Font(name = "Arial Black",size =11,color=Light_Blue)
    ws[f"{concat2}2"].font = Font(name = "Arial Black",size =11,color=Light_Blue)
    val = 3
    for count in myList:
        cel = f'{cell2}{str(val)}'
        ws[cel].value  = "Absent"
        white = "FFFFFFFF"
        ws[cel].alignment = Alignment(horizontal='center', vertical='center')
        ws[cel].font = Font(name="Tahoma", size=10, color=white)
        ws[cel].fill = PatternFill(start_color=light_red, end_color=light_red,fill_type = "solid")
        thin = Side(border_style="thin", color=pink)
        double = Side(border_style="double", color=green)
        ws[cel].border = Border(top=double, left=thin, right=thin, bottom=double)
        val+=1

# This function reads the previously stored data from the "Extra_Data.csv" file
def Read_CSV():
    with open("C:\\Users\\Administrator\\Documents\\FaceRecognitionProject\\RootFolder\\Extra_Data.csv") as file_obj:
        reader_obj = csv.reader(file_obj)
        for row in reader_obj:
            rw = row
    file_obj.close()
    return rw

# This function writes the new data to the "Extra_Data.csv" file
def Write_CSV(Col_Value1, Col_Value2, Col_nmr):
    f = open("C:\\Users\\Administrator\\Documents\\FaceRecognitionProject\\RootFolder\\Extra_Data.csv","w")
    col_nr = int(Col_nmr) + 2
    today = date.today()
    d1 = today.strftime("%d/%m/%y")
    # d1 = "09/01/22"
    f.write(f'{d1},{Col_Value1},{Col_Value2},{col_nr}')
    f.close()

# This function checks whether the Date store in the "Extra_Data.csv" file and the current date are similar or not
def Check_Date_Similarity():
    today = date.today()
    d1 = today.strftime("%d/%m/%y")
    # d1 = "09/01/22"
    d2 = Read_CSV()[0]
    if(d1 == d2):
        return True
    else:
        return False

# This function generates the column name for the new register to be created every time the date changes 
def Generate_Column_Name(Alp):
    length = len(Alp)
    To_int = ord(Alp[length-1])
    
    if(length>1 and Alp[length-1] != 'Z' and Alp != "YY"):
        val1 = f"{Alp[length-2]}{chr(To_int +1)}"
        val2 = f"{Alp[length-2]}{chr(To_int +2)}"
        return val1,val2
    elif(length>1 and Alp[length-1] == 'Z' and Alp !="YY"):
        To_int2= ord(Alp[length-2] )
        val1 = f"{chr(To_int2 +1)}A"
        val2 = f"{chr(To_int2 +1)}B"
        return val1,val2
    elif(length>1 and Alp =="YY"):
        return "YZ","ZA"
    else:
        vl1 = chr(To_int +1)
        vl2 = chr(To_int +2)
        return vl1, vl2 

# This function marks the attendance in the excel file if the face matches to our database images 
def MarkAttendance(name,col1_val,col2_val):
    i = 1
    while(i<=file_count):
        name_cel = f"A{i+2}"
        time_cel = f"{col1_val}{i+2}"
        status_cel  =f"{col2_val}{i+2}"
        if(ws[name_cel].value.upper() == name and ws[status_cel].value =="Absent"):
            now = datetime.now()
            dtString = now.strftime("%H:%M:%S")
            ws[time_cel].value = dtString
            ws[time_cel].alignment = Alignment(horizontal='center', vertical='center')
            ws[status_cel].value = "Present"
            ws[status_cel].alignment = Alignment(horizontal='center', vertical='center')
            ws[status_cel].font = Font(name="Tahoma", size=10, color=white)
            ws[status_cel].fill = PatternFill(start_color=light_green, end_color=light_green,fill_type = "solid")
        i=i+1
    wb.save(path + "Attendance_Final.xlsx")


# Function Calls 
Setup()
c1 = Read_CSV()
v1 = Generate_Column_Name(c1[2])
if (Check_Date_Similarity() == False):
    v1 = Generate_Column_Name(c1[2])
    Create(v1[0],v1[1],int(c1[3]))
    Write_CSV(v1[0],v1[1],c1[3])
    c1 = Read_CSV()
    v = Read_CSV()[2]

print("\nEncoding images and activating modules please wait .......")
encodeListKnown = findEncodings(images)
print("Encoding complete ")


cap = cv2.VideoCapture(0)
while True:
    success ,img = cap.read() 
    imgS = cv2.resize(img,(0,0),None,0.25,0.25)
    imgS = cv2.cvtColor(imgS,cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodeCurFrame = face_recognition.face_encodings(imgS,facesCurFrame)

    # comparing the images 
    for encodeFace, faceLoc in zip(encodeCurFrame,facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown,encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown,encodeFace)

        matchIndex = np.argmin(faceDis)
        if matches[matchIndex]:
            name = classNames[matchIndex].upper()
            print(name)
            y1,x2,y2,x1 = faceLoc
            y1,x2,y2,x1 = y1*4,x2*4,y2*4,x1*4

            cv2.rectangle(img,(x1,y1),(x2,y2),(0,255,0),2)
            cv2.rectangle(img,(x1,y2 -35),(x2,y2),(0,255,0),cv2.FILLED)
            cv2.putText(img,name,(x1+6,y2-6),cv2.FONT_HERSHEY_COMPLEX,1,(255,255,255),2)
            MarkAttendance(name,v1[0],v1[1])
    cv2.imshow("WebCam",img)
    cv2.waitKey(1)


    
